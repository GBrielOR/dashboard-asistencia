import io
import streamlit as st
import pandas as pd
import plotly.express as px
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from fpdf import FPDF, XPos, YPos
from datetime import datetime

st.set_page_config(page_title="Control de Asistencias - Unidad Educativa Caranqui", layout="wide")

st.title("Dashboard de Asistencias Escolares")
st.markdown("### Unidad Educativa Caranqui - Periodo Lectivo 2025 - 2026")

ARCHIVO_EXCEL = "Asistencia_Escolar_CORREGIDA_FINAL.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS: generar imágenes PNG con matplotlib (sin Kaleido)
# ─────────────────────────────────────────────────────────────────────────────

def _grafico_pastel_png(df):
    """Genera el gráfico de pastel como PNG en memoria usando matplotlib."""
    conteo = df['Estado_Nombre'].value_counts()
    colores_map = {
        'Presente':       '#2ecc71',
        'Ausente (Falta)':'#e74c3c',
        'Tarde (Atraso)': '#f1c40f',
        'Justificado':    '#3498db',
    }
    colores = [colores_map.get(k, '#aaaaaa') for k in conteo.index]

    fig, ax = plt.subplots(figsize=(5, 3.5), facecolor='white')
    wedges, texts, autotexts = ax.pie(
        conteo.values,
        labels=None,
        colors=colores,
        autopct='%1.1f%%',
        startangle=90,
        wedgeprops=dict(width=0.6),   # donut
        pctdistance=0.75
    )
    for at in autotexts:
        at.set_fontsize(8)
    ax.legend(
        wedges, conteo.index,
        loc="lower center", bbox_to_anchor=(0.5, -0.18),
        ncol=2, fontsize=7, frameon=False
    )
    ax.set_title("Distribución General de Asistencias", fontsize=10, fontweight='bold', pad=8)
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


def _grafico_barras_png(df):
    """Genera el gráfico de barras por curso como PNG en memoria usando matplotlib."""
    asist = (
        df.groupby(['Curso', 'Estado'])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for c in ['P', 'T', 'A', 'J']:
        if c not in asist.columns:
            asist[c] = 0
    asist['Total'] = asist[['P', 'T', 'A', 'J']].sum(axis=1)
    asist['pct'] = ((asist['P'] + asist['T']) / asist['Total']) * 100

    fig, ax = plt.subplots(figsize=(5, 3.5), facecolor='white')
    cursos = asist['Curso'].tolist()
    valores = asist['pct'].tolist()
    colores = plt.cm.tab10.colors[:len(cursos)]

    bars = ax.bar(cursos, valores, color=colores, edgecolor='white', linewidth=0.5)
    ax.set_ylim(60, 101)
    ax.set_ylabel('Porcentaje (%)', fontsize=8)
    ax.set_title('Porcentaje de Asistencia por Grado', fontsize=10, fontweight='bold', pad=8)
    ax.tick_params(axis='x', labelsize=7, rotation=30)
    ax.tick_params(axis='y', labelsize=7)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    for bar, val in zip(bars, valores):
        ax.text(bar.get_x() + bar.get_width() / 2, val + 0.3,
                f'{val:.1f}%', ha='center', va='bottom', fontsize=7)
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data
def procesar_todo_el_excel(ruta_archivo):
    xl = pd.ExcelFile(ruta_archivo)
    hojas = xl.sheet_names

    lista_registros = []
    meses_dict = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
        "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
    }

    progreso_texto = st.empty()

    for hoja in hojas:
        hoja_limpia = hoja.strip()
        if hoja_limpia.lower() in ["inicio", ""] or len(hoja_limpia) < 2:
            continue

        progreso_texto.text(f"Cargando datos: {hoja_limpia}")

        año = "2026"
        if "2025" in hoja_limpia:
            año = "2025"
        elif "2026" in hoja_limpia:
            año = "2026"

        curso = hoja_limpia.replace("2025", "").replace("2026", "").replace("-", "").strip()
        if curso == "":
            curso = hoja_limpia

        try:
            df_hoja = pd.read_excel(ruta_archivo, sheet_name=hoja, header=None)
        except Exception:
            continue

        mes_actual = None

        for idx, row in df_hoja.iterrows():
            valores_fila = [str(x).strip().upper() for x in row.values if pd.notna(x)]

            if not valores_fila:
                continue

            texto_unido = " ".join(valores_fila)
            if "ASISTENCIA" in texto_unido and "AÑO" in texto_unido:
                continue

            encontro_mes = False
            for palabra in valores_fila:
                if palabra in meses_dict:
                    mes_actual = palabra
                    encontro_mes = True
                    break
            if encontro_mes:
                continue

            if "ESTUDIANTE" in texto_unido or "N°" in texto_unido or "PRES." in texto_unido:
                continue

            estudiante = None
            col_estudiante_idx = -1

            for i, celda in enumerate(row.values):
                celda_str = str(celda).strip()
                if pd.notna(celda) and celda_str != "" and not celda_str.replace(".", "").isdigit():
                    if "TOTAL" in celda_str.upper() or "ESTUDIANTE" in celda_str.upper():
                        break
                    estudiante = celda_str
                    col_estudiante_idx = i
                    break

            if estudiante and mes_actual:
                for dia in range(1, 32):
                    col_idx = col_estudiante_idx + dia
                    if col_idx < len(row):
                        estado = row.iloc[col_idx]

                        if pd.isna(estado) or str(estado).strip() in ["", "0", "0.0"]:
                            estado_final = "P"
                        else:
                            estado_final = str(estado).strip().upper()

                        if estado_final in ['P', 'A', 'T', 'J']:
                            lista_registros.append({
                                "Año": año,
                                "Curso": curso,
                                "Estudiante": estudiante,
                                "Mes": mes_actual,
                                "Dia": dia,
                                "Estado": estado_final
                            })

    progreso_texto.empty()

    if len(lista_registros) == 0:
        return pd.DataFrame()

    df_final = pd.DataFrame(lista_registros)
    mapa_estados = {'P': 'Presente', 'A': 'Ausente (Falta)', 'T': 'Tarde (Atraso)', 'J': 'Justificado'}
    df_final['Estado_Nombre'] = df_final['Estado'].map(mapa_estados)
    return df_final


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DEL PDF
# ─────────────────────────────────────────────────────────────────────────────

def _cell(pdf, w, h, txt, bold=False, align="L", border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT):
    """Wrapper para pdf.cell con nueva API de fpdf2."""
    pdf.cell(w, h, txt, border=border, align=align, new_x=new_x, new_y=new_y)

def generar_reporte_pdf(año_sel, curso_sel, t_asistencia, t_faltas, t_atrasos,
                       df_ranking, img_pastel=None, img_barra=None):
    pdf = FPDF()
    
    # ─────────────────────────────────────────────────────────────────────────
    # PÁGINA 1: RESUMEN ESTADÍSTICO Y ALERTA TEMPRANA (TABLA)
    # ─────────────────────────────────────────────────────────────────────────
    pdf.add_page()

    # Encabezado institucional
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "UNIDAD EDUCATIVA CARANQUI",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")

    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, f"Reporte de Control de Asistencias - Periodo {año_sel}",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.cell(0, 6, f"Filtro aplicado: {curso_sel}",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.cell(0, 6, f"Fecha de emision: {datetime.now().strftime('%d/%m/%Y')}",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(8)

    # 1. Resumen estadístico
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "1. RESUMEN ESTADISTICO GENERAL",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    datos_kpi = [
        ("Tasa de Asistencia Promedio:", f"{t_asistencia:.1f}%"),
        ("Total Inasistencias (Faltas):", f"{t_faltas:,}"),
        ("Total de Atrasos Registrados:", f"{t_atrasos:,}"),
    ]
    for label, valor in datos_kpi:
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(90, 7, label, new_x=XPos.RIGHT, new_y=YPos.LAST)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, valor, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    # 2. Alerta temprana (Ahora va inmediatamente debajo de los KPIs)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "2. ALERTA TEMPRANA: ESTUDIANTES CON MAYOR AUSENTISMO",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    if 'Ausente (Falta)' in df_ranking.columns and len(df_ranking) > 0:
        # Encabezados de la tabla
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(40, 7, "Curso",         border=1, align="C", new_x=XPos.RIGHT, new_y=YPos.LAST)
        pdf.cell(110, 7, "Estudiante",   border=1, align="L", new_x=XPos.RIGHT, new_y=YPos.LAST)
        pdf.cell(40, 7, "Faltas Injust.",border=1, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # Contenido limitado a los primeros 12-15 alumnos para asegurar que entre en la hoja 1
        pdf.set_font("Helvetica", "", 9)
        ranking_top = df_ranking.sort_values(by='Ausente (Falta)', ascending=False).head(13)
        for _, fila in ranking_top.iterrows():
            pdf.cell(40,  6, str(fila['Curso']),                    border=1, align="C", new_x=XPos.RIGHT,  new_y=YPos.LAST)
            pdf.cell(110, 6, str(fila['Estudiante'])[:55],          border=1, align="L", new_x=XPos.RIGHT,  new_y=YPos.LAST)
            pdf.cell(40,  6, str(int(fila['Ausente (Falta)'])),     border=1, align="C", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    else:
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 7, "No se registran faltas en el grupo seleccionado.",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # ─────────────────────────────────────────────────────────────────────────
    # PÁGINA 2: GRÁFICOS Y FIRMAS
    # ─────────────────────────────────────────────────────────────────────────
    if img_pastel is not None or img_barra is not None:
        pdf.add_page()  # Enviamos los gráficos limpios a la siguiente página

        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "3. GRAFICOS DE ASISTENCIA",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(6)

        ancho_grafico = 92   # mm
        alto_grafico = 65    # mm
        y_antes = pdf.get_y()

        if img_pastel:
            img_pastel.seek(0)
            pdf.image(img_pastel, x=10, y=y_antes, w=ancho_grafico, h=alto_grafico)

        if img_barra:
            img_barra.seek(0)
            pdf.image(img_barra, x=108, y=y_antes, w=ancho_grafico, h=alto_grafico)

        # Movemos el cursor debajo de los gráficos para estampar las firmas
        pdf.set_y(y_antes + alto_grafico + 20)

    else:
        # En caso de que no existan imágenes por algún error, dejamos las firmas en la pág 1
        pdf.ln(15)

    # Bloque de Firmas (Ahora queda fijo en la parte inferior de la página de gráficos)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(95, 8, "_________________________", new_x=XPos.RIGHT, new_y=YPos.LAST, align="C")
    pdf.cell(95, 8, "_________________________", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.cell(95, 5, "Inspector / Responsable",   new_x=XPos.RIGHT, new_y=YPos.LAST, align="C")
    pdf.cell(95, 5, "Rectorado / Direccion",     new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")

    return bytes(pdf.output())
# ─────────────────────────────────────────────────────────────────────────────
# APLICACIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

try:
    df_asistencias = procesar_todo_el_excel(ARCHIVO_EXCEL)

    if df_asistencias.empty:
        st.error("No se pudieron extraer registros. Compruebe que el archivo "
                 "'Asistencia_Escolar_CORREGIDA_FINAL.xlsx' no esté vacío o abierto en Excel.")
    else:
        st.sidebar.header("Filtros de Control")

        lista_años = sorted(df_asistencias['Año'].unique())
        filtro_año = st.sidebar.selectbox("Seleccione el Año Lectivo:", options=lista_años)
        df_filtrado_año = df_asistencias[df_asistencias['Año'] == filtro_año]

        lista_cursos = sorted(df_filtrado_año['Curso'].unique())
        filtro_curso = st.sidebar.selectbox(
            "Seleccione el Curso / Grado:",
            options=["Todos los Cursos"] + lista_cursos
        )

        df_final_display = (
            df_filtrado_año[df_filtrado_año['Curso'] == filtro_curso]
            if filtro_curso != "Todos los Cursos"
            else df_filtrado_año
        )

        total_registros = len(df_final_display)
        if total_registros > 0:
            asistencias_ok     = len(df_final_display[df_final_display['Estado'].isin(['P', 'T'])])
            porcentaje_asist   = (asistencias_ok / total_registros) * 100
            total_faltas       = len(df_final_display[df_final_display['Estado'] == 'A'])
            total_atrasos      = len(df_final_display[df_final_display['Estado'] == 'T'])
            ranking_calculo    = (
                df_final_display
                .groupby(['Curso', 'Estudiante', 'Estado_Nombre'])
                .size()
                .unstack(fill_value=0)
                .reset_index()
            )

            # ── Figuras Plotly para la UI ──────────────────────────────────
            fig_pastel = px.pie(
                df_final_display, names='Estado_Nombre', color='Estado_Nombre',
                color_discrete_map={
                    'Presente':       '#2ecc71',
                    'Ausente (Falta)':'#e74c3c',
                    'Tarde (Atraso)': '#f1c40f',
                    'Justificado':    '#3498db'
                }, hole=0.4
            )

            asist_curso = (
                df_final_display.groupby(['Curso', 'Estado'])
                .size().unstack(fill_value=0).reset_index()
            )
            for c in ['P', 'T', 'A', 'J']:
                if c not in asist_curso.columns:
                    asist_curso[c] = 0
            asist_curso['Total'] = asist_curso[['P', 'T', 'A', 'J']].sum(axis=1)
            asist_curso['% Asistencia'] = (
                (asist_curso['P'] + asist_curso['T']) / asist_curso['Total']
            ) * 100

            fig_barra = px.bar(
                asist_curso, x='Curso', y='% Asistencia',
                title="Porcentaje de Asistencia por Grado",
                labels={'% Asistencia': 'Porcentaje (%)'}, color='Curso'
            )
            fig_barra.update_yaxes(range=[60, 101])

            # ── Imágenes matplotlib para el PDF (sin Kaleido) ─────────────
            img_pastel_buf = _grafico_pastel_png(df_final_display)
            img_barra_buf  = _grafico_barras_png(df_final_display)

            # ── Generar PDF ────────────────────────────────────────────────
            pdf_bytes = generar_reporte_pdf(
                filtro_año, filtro_curso,
                porcentaje_asist, total_faltas, total_atrasos, ranking_calculo,
                img_pastel=img_pastel_buf,
                img_barra=img_barra_buf
            )

            st.sidebar.markdown("---")
            st.sidebar.subheader("Exportar Documento")
            st.sidebar.download_button(
                label="Descargar PDF para Imprimir",
                data=pdf_bytes,
                file_name=f"Reporte_Asistencia_{filtro_curso.replace(' ', '_')}_{filtro_año}.pdf",
                mime="application/pdf"
            )

            # ── KPIs ──────────────────────────────────────────────────────
            kpi1, kpi2, kpi3 = st.columns(3)
            kpi1.metric("Tasa de Asistencia Promedio", f"{porcentaje_asist:.1f}%")
            kpi2.metric("Total Inasistencias (Faltas)", f"{total_faltas:,}")
            kpi3.metric("Total de Atrasos Registrados", f"{total_atrasos:,}")

            st.markdown("---")

            # ── Gráficos interactivos ──────────────────────────────────────
            col_izq, col_der = st.columns(2)
            with col_izq:
                st.subheader("Distribución General de Asistencias")
                st.plotly_chart(fig_pastel, use_container_width=True)
            with col_der:
                st.subheader("Comparativa de Asistencia por Cursos")
                st.plotly_chart(fig_barra, use_container_width=True)

            st.markdown("---")

            # ── Tabla de alerta temprana ───────────────────────────────────
            st.subheader("Alerta Temprana: Estudiantes con Mayor Cantidad de Faltas")
            if 'Ausente (Falta)' in ranking_calculo.columns:
                ranking_display = (
                    ranking_calculo
                    .sort_values(by='Ausente (Falta)', ascending=False)
                    .head(15)
                    .rename(columns={'Ausente (Falta)': 'Faltas Injustificadas'})
                )
                st.dataframe(ranking_display[['Curso', 'Estudiante', 'Faltas Injustificadas']],
                             use_container_width=True)
            else:
                st.info("No se registran inasistencias en este grupo.")
        else:
            st.warning("No hay datos para mostrar con los filtros seleccionados.")

except FileNotFoundError:
    st.error(f"No se encontró el archivo '{ARCHIVO_EXCEL}' en la carpeta especificada.")
except Exception as e:
    st.error(f"Ocurrió un error inesperado al procesar los datos: {e}")